import openpyxl
from web3 import Web3

workbook = openpyxl.load_workbook("wallets.xlsx")

worksheet = workbook.active

addresses = [str(cell.value) for cell in worksheet["A"][1:] if cell.value]

node_url = "your_rpc"
token_address = Web3.to_checksum_address("0x...")

w3 = Web3(Web3.HTTPProvider(node_url))

with open("erc20.json") as f:
    abi = f.read()
contract = w3.eth.contract(address=token_address, abi=abi)


def GetAccountBalances(addresses):
    results = []

    for address in addresses:
        balance = contract.functions.balanceOf(Web3.to_checksum_address(address)).call()
        balance = balance // 10 ** 18
        results.append(balance)
        print(f"{address}: {balance}")

    return results


if __name__ == '__main__':
    results = GetAccountBalances(addresses)
    print(results)
    total_count = 0
    for row, (address, tx_count) in enumerate(zip(addresses, results), start=2):
        worksheet.cell(row=row, column=2, value=tx_count)
        total_count += tx_count
    worksheet.cell(row=len(addresses)+2, column=1, value="Total count")
    worksheet.cell(row=len(addresses)+2, column=2, value=total_count)
    workbook.save("wallets.xlsx")
